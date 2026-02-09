import os
import re
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from werkzeug.utils import secure_filename
from flask import Flask, request, render_template, send_file, redirect, url_for
from flask_cors import CORS

# Initialize Flask app
app = Flask(__name__)
CORS(app)

# File upload configurations
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['ALLOWED_EXTENSIONS'] = {'xls', 'xlsx'}

# Helper function to check file extensions
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DRAFT_TEMPLATE_PATH = os.path.join(BASE_DIR, "Draft Output.xlsx")

def extract_no_faktur_from_description(desc: str) -> str | None:
    if pd.isna(desc):
        return None
    s = str(desc).strip()
    parts = [p.strip() for p in s.split("/")]
    # "slash ke-2" = index 1 (setelah slash pertama)
    return parts[1] if len(parts) >= 2 and parts[1] else None


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalisasi nama kolom: uppercase, spasi/punctuation -> underscore, rapihin underscore.
    Contoh: 'No Voucher' -> 'NO_VOUCHER'
    """
    def clean(col):
        col = str(col).strip().upper()
        col = re.sub(r"[^A-Z0-9]+", "_", col)
        col = re.sub(r"_+", "_", col).strip("_")
        return col

    df = df.copy()
    df.columns = [clean(c) for c in df.columns]
    return df

def _parse_id_number(x):
    """
    Aman untuk angka dengan format Indonesia:
    - '77.597.727' -> 77597727
    - '7.759.773'  -> 7759773
    - '1.234,56'   -> 1234.56
    """
    if pd.isna(x):
        return np.nan
    if isinstance(x, (int, float, np.number)):
        return float(x)

    s = str(x).strip()
    if not s:
        return np.nan

    s = s.replace(" ", "")
    # hapus ribuan ".", ubah desimal "," jadi "."
    s = s.replace(".", "")
    s = s.replace(",", ".")
    # handle (123) -> -123
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]

    try:
        v = float(s)
        return -v if neg else v
    except:
        return np.nan
    
def compare_files(k3_path: str, coretax_path_1: str, coretax_path_2: str, output_dir: str) -> str:
    # 1) Read files
    k3 = pd.read_excel(k3_path, header=1)
    coretax_1 = pd.read_excel(coretax_path_1, header=1)  # FP Digunggung
    coretax_2 = pd.read_excel(coretax_path_2, header=1)  # FP Tidak Digunggung

    # 2) Normalize columns for Coretax (biar NO VOUCHER / DOC_NO kebaca konsisten)
    coretax_1 = _normalize_columns(coretax_1)
    coretax_2 = _normalize_columns(coretax_2)

    # Pastikan key jadi NO_VOUCHER
    if "DOC_NO" in coretax_1.columns and "NO_VOUCHER" not in coretax_1.columns:
        coretax_1 = coretax_1.rename(columns={"DOC_NO": "NO_VOUCHER"})
    if "DOC_NO" in coretax_2.columns and "NO_VOUCHER" not in coretax_2.columns:
        coretax_2 = coretax_2.rename(columns={"DOC_NO": "NO_VOUCHER"})

    if "NO_VOUCHER" not in coretax_1.columns:
        raise ValueError("Coretax Digunggung: kolom DOC_NO / NO VOUCHER tidak ditemukan.")
    if "NO_VOUCHER" not in coretax_2.columns:
        raise ValueError("Coretax Tidak Digunggung: kolom DOC_NO / NO VOUCHER tidak ditemukan.")

    # 3) Harmonize DPP/PPN + CUSTOMER + status
    # --- Digunggung: AMOUNT_BEF_TAX = DPP, TAX_AMOUNT = PPN, CUSTOMER_NAME = CUSTOMER
    if "DPP" not in coretax_1.columns and "AMOUNT_BEF_TAX" in coretax_1.columns:
        coretax_1["DPP"] = coretax_1["AMOUNT_BEF_TAX"]
    if "PPN" not in coretax_1.columns and "TAX_AMOUNT" in coretax_1.columns:
        coretax_1["PPN"] = coretax_1["TAX_AMOUNT"]

    coretax_1["CUSTOMER"] = coretax_1["CUSTOMER_NAME"] if "CUSTOMER_NAME" in coretax_1.columns else None
    coretax_1["FP_STATUS"] = "FP Digunggung"

    # --- Tidak Digunggung: DPP = DPP, PPN = PPN, NAMA_PEMBELI = CUSTOMER
    if "DPP" not in coretax_2.columns and "AMOUNT_BEF_TAX" in coretax_2.columns:
        coretax_2["DPP"] = coretax_2["AMOUNT_BEF_TAX"]
    if "PPN" not in coretax_2.columns and "TAX_AMOUNT" in coretax_2.columns:
        coretax_2["PPN"] = coretax_2["TAX_AMOUNT"]

    if "NAMA_PEMBELI" in coretax_2.columns:
        coretax_2["CUSTOMER"] = coretax_2["NAMA_PEMBELI"]
    elif "CUSTOMER_NAME" in coretax_2.columns:
        coretax_2["CUSTOMER"] = coretax_2["CUSTOMER_NAME"]
    else:
        coretax_2["CUSTOMER"] = None

    coretax_2["FP_STATUS"] = "FP Tidak Digunggung"

    # 4) Bersihin key + convert angka
    for df in (coretax_1, coretax_2):
        df["NO_VOUCHER"] = df["NO_VOUCHER"].astype(str).str.strip()
        if "DPP" in df.columns:
            df["DPP"] = df["DPP"].apply(_parse_id_number)
        else:
            df["DPP"] = 0.0
        if "PPN" in df.columns:
            df["PPN"] = df["PPN"].apply(_parse_id_number)
        else:
            df["PPN"] = 0.0

    # 5) Combine Coretax (ambil kolom penting aja)
    keep_cols_1 = [c for c in ["NO_VOUCHER", "VOUCHER_NO", "DPP", "PPN", "CUSTOMER", "FP_STATUS"] if c in coretax_1.columns]
    keep_cols_2 = [c for c in ["NO_VOUCHER", "VOUCHER_NO", "DPP", "PPN", "CUSTOMER", "FP_STATUS"] if c in coretax_2.columns]
    coretax_combined = pd.concat([coretax_1[keep_cols_1], coretax_2[keep_cols_2]], ignore_index=True)

    # 6) kalau NO_VOUCHER muncul beberapa kali, DPP/PPN dijumlah, CUSTOMER diambil first non-null, status digabung unik
    def join_unique(series):
        vals = [v for v in series.dropna().astype(str).tolist() if v.strip()]
        return "; ".join(sorted(set(vals))) if vals else None

    agg_map = {
        "DPP": "sum",
        "PPN": "sum",
        "CUSTOMER": "first",
        "FP_STATUS": join_unique
    }
    if "VOUCHER_NO" in coretax_combined.columns:
        agg_map["VOUCHER_NO"] = "first"

    # Aggregate the combined coretax data
    coretax_agg = coretax_combined.groupby("NO_VOUCHER", as_index=False).agg(agg_map)

    # 7) Extract no faktur dari Description
    k3["No Faktur (key)"] = k3["Description"].apply(extract_no_faktur_from_description)
    k3["No Faktur (key)"] = k3["No Faktur (key)"].astype(str).str.strip()

    # 8) Debugging step: Check columns in Coretax_2
    print("Columns in Coretax_2:", coretax_2.columns)

    # 9) Debugging: Check if 'NO FP MODIF' exists in coretax_2
    if "NO FP MODIF" in coretax_2.columns:
        print("NO FP MODIF exists in Coretax_2.")
    else:
        print("NO FP MODIF NOT found in Coretax_2")

    # 10) Merge
    merged = pd.merge(
        k3,
        coretax_agg,
        left_on="No Faktur (key)",
        right_on="NO_VOUCHER",
        how="left",
        indicator=True
    )

    # 11) Compute difference (tetap pakai NET seperti versi kamu)
    merged["Debit Amount"] = pd.to_numeric(merged["Debit Amount"], errors="coerce").fillna(0)
    merged["Credit Amount"] = pd.to_numeric(merged["Credit Amount"], errors="coerce").fillna(0)
    merged["K3_NET"] = merged["Debit Amount"] - merged["Credit Amount"]

    merged["DPP"] = pd.to_numeric(merged["DPP"], errors="coerce").fillna(0)
    merged["PPN"] = pd.to_numeric(merged["PPN"], errors="coerce").fillna(0)
    merged["Difference"] = merged["K3_NET"] - (merged["DPP"] + merged["PPN"])

    # 12) Keterangan + Customer (langsung dari kolom kanonik)
    merged["Keterangan (Digunggung/Tidak Digunngung)"] = merged["FP_STATUS"]
    merged.loc[merged["_merge"] != "both", "Keterangan (Digunggung/Tidak Digunngung)"] = "Tidak ada di Coretax"

    merged["Customer"] = merged["CUSTOMER"]
    merged.loc[merged["_merge"] != "both", "Customer"] = None

    # Debugging: Check if 'NO_FP_MODIF' exists in coretax_2
    print("Cek apakah 'NO_FP_MODIF' ada di coretax_2:", "NO_FP_MODIF" in coretax_2.columns)
    if "NO_FP_MODIF" in coretax_2.columns:
        print(coretax_2["NO_FP_MODIF"].head())  # Menampilkan beberapa nilai untuk memastikan kolom ada
    else:
        print("Kolom 'NO_FP_MODIF' tidak ditemukan di Coretax_2")

    print("Setelah merge, kolom di merged:", merged.columns)

    # Jika kolom 'NO_FP_MODIF' ada di coretax_2, tambahkan ke merged
    if "NO_FP_MODIF" in coretax_2.columns:
        merged = pd.merge(
            merged,
            coretax_2[["NO_VOUCHER", "NO_FP_MODIF"]],
            left_on="No Faktur (key)",
            right_on="NO_VOUCHER",
            how="left",
            suffixes=("", "_from_coretax2")
        )
        print("Setelah merge NO_FP_MODIF, kolom di merged:", merged.columns)
    else:
        merged["NO_FP_MODIF"] = None  # Atur sebagai None jika kolom tidak ada

    # 13) Before filling NaN, convert categorical columns to string type
    for column in merged.columns:
        if merged[column].dtype.name == 'category':  # Check if the column is categorical
            merged[column] = merged[column].astype(str)
    # Now, proceed with other operations
    merged.fillna("-", inplace=True)

    # 14) Write ke template
    if not os.path.exists(DRAFT_TEMPLATE_PATH):
        raise FileNotFoundError("Draft Output.xlsx tidak ditemukan. Taruh file itu 1 folder dengan app.py")

    wb = load_workbook(DRAFT_TEMPLATE_PATH)
    ws = wb.active

    start_row = 5
    max_row = ws.max_row
    for r in range(start_row, max_row + 1):
        for c in range(1, 19):
            ws.cell(r, c).value = None

    for i in range(len(merged)):
        r = start_row + i
        row = merged.iloc[i]

        # Left (K3)
        ws.cell(r, 1).value = row.get("Account No")
        ws.cell(r, 2).value = row.get("Account Name")
        ws.cell(r, 3).value = row.get("Date")
        ws.cell(r, 4).value = row.get("Voucher Category")
        ws.cell(r, 5).value = row.get("Voucher No.")
        ws.cell(r, 6).value = row.get("Description")
        ws.cell(r, 7).value = row.get("Debit Amount")
        ws.cell(r, 8).value = row.get("Credit Amount")
        ws.cell(r, 9).value = row.get("Direction")
        ws.cell(r, 10).value = row.get("Balance")

        # Right (Coretax)
        ws.cell(r, 12).value = row.get("NO_VOUCHER")  # No Faktur from Coretax
        ws.cell(r, 13).value = row.get("NO_FP_MODIF")  # Voucher No. from NO FP MODIF
        ws.cell(r, 14).value = row.get("DPP")
        ws.cell(r, 15).value = row.get("PPN")
        ws.cell(r, 16).value = row.get("Difference")
        ws.cell(r, 17).value = row.get("Customer")
        ws.cell(r, 18).value = row.get("Keterangan (Digunggung/Tidak Digunngung)")

    os.makedirs(output_dir, exist_ok=True)
    out_name = f"Draft_Updated_Output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = os.path.join(output_dir, out_name)
    wb.save(out_path)
    return out_path

# Fungsi untuk menghapus file output
def delete_output_files(output_dir):
    try:
        # Cek jika direktori output ada
        if os.path.exists(output_dir):
            # Hapus semua file dalam folder output
            for filename in os.listdir(output_dir):
                file_path = os.path.join(output_dir, filename)
                if os.path.isfile(file_path):
                    os.remove(file_path)  # Hapus file
                    print(f"File deleted: {file_path}")
    except Exception as e:
        print(f"Error deleting files: {e}")

# Homepage route to upload files
@app.route('/')
def home():
    return render_template('index.html')

# Handle file upload and processing
@app.route('/upload', methods=['POST'])
def upload_file():
    if 'k3_file' not in request.files or 'coretax_file_1' not in request.files or 'coretax_file_2' not in request.files:
        return 'No file part'
    
    k3_file = request.files['k3_file']
    coretax_file_1 = request.files['coretax_file_1']
    coretax_file_2 = request.files['coretax_file_2']

    # Check if files have been selected
    if k3_file.filename == '' or coretax_file_1.filename == '' or coretax_file_2.filename == '':
        return 'No selected file'

    # Check if files are valid
    if k3_file and allowed_file(k3_file.filename) and coretax_file_1 and allowed_file(coretax_file_1.filename) and coretax_file_2 and allowed_file(coretax_file_2.filename):
        k3_filename = secure_filename(k3_file.filename)
        coretax_filename_1 = secure_filename(coretax_file_1.filename)
        coretax_filename_2 = secure_filename(coretax_file_2.filename)

        # Save the uploaded files in the correct directory
        k3_file_path = os.path.join(app.config['UPLOAD_FOLDER'], k3_filename)
        coretax_file_path_1 = os.path.join(app.config['UPLOAD_FOLDER'], coretax_filename_1)
        coretax_file_path_2 = os.path.join(app.config['UPLOAD_FOLDER'], coretax_filename_2)

        k3_file.save(k3_file_path)
        coretax_file_1.save(coretax_file_path_1)
        coretax_file_2.save(coretax_file_path_2)

        # Define output directory for the draft output
        output_dir = os.path.join(BASE_DIR, "outputs")
        os.makedirs(output_dir, exist_ok=True)

        # Compare files and update draft output
        updated_file = compare_files(k3_file_path, coretax_file_path_1, coretax_file_path_2, output_dir)

        # Cleanup: Delete uploaded files after processing
        delete_all_uploaded_files()

        # Redirect to comparison route with updated file and page number
        return redirect(url_for('show_comparison', updated_file=updated_file, page=1))

    return 'Invalid file type'


@app.route('/comparison', methods=['GET'])
def show_comparison():
    updated_file = request.args.get('updated_file')
    page = request.args.get('page', 1, type=int)  # Default to 1 if 'page' is not in the URL
    rows_per_page = 6

    # Read the updated file for comparison
    merged_df = pd.read_excel(updated_file)

    # Pagination Logic
    total_pages = (len(merged_df) // rows_per_page) + (1 if len(merged_df) % rows_per_page != 0 else 0)
    start_row = (page - 1) * rows_per_page
    end_row = start_row + rows_per_page
    page_data = merged_df[start_row:end_row]

    table_html = page_data.to_html(classes='table table-striped')

    # Pagination controls (Previous/Next buttons)
    pagination_html = ""
    if page > 1:
        pagination_html += f'<a href="/comparison?updated_file={updated_file}&page={page-1}" class="btn btn-secondary">Previous</a>'

    # Display numbers for pagination, limit to a range of 5 numbers
    # Ensure that the pagination numbers are in a range that doesn't exceed the total pages
    pagination_start = max(1, page - 2)  # Start from 2 pages before the current page
    pagination_end = min(total_pages, pagination_start + 4)  # End at 5 pages from the starting page

    for p in range(pagination_start, pagination_end + 1):
        if p == page:
            pagination_html += f' <span class="btn btn-light disabled">{p}</span>'
        else:
            pagination_html += f' <a href="/comparison?updated_file={updated_file}&page={p}" class="btn btn-secondary">{p}</a>'

    if page < total_pages:
        pagination_html += f' <a href="/comparison?updated_file={updated_file}&page={page+1}" class="btn btn-secondary">Next</a>'

    # Return the table with pagination controls, passing `page` to the template
    return render_template('comparison.html', table_html=table_html, pagination_html=pagination_html, updated_file=updated_file, page=page)


@app.route('/download/<filename>')
def download_file(filename):
    return send_file(os.path.join(BASE_DIR, 'outputs', filename), as_attachment=True)

@app.route('/clear_outputs', methods=['GET'])
def clear_outputs():
    output_dir = os.path.join(BASE_DIR, "outputs")
    delete_output_files(output_dir)
    return redirect(url_for('home'))  # Kembali ke halaman utama

def delete_all_uploaded_files():
    try:
        # Cek jika direktori upload ada
        if os.path.exists(app.config['UPLOAD_FOLDER']):
            # Hapus semua file dalam folder upload
            for filename in os.listdir(app.config['UPLOAD_FOLDER']):
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                if os.path.isfile(file_path):
                    os.remove(file_path)  # Hapus file
                    print(f"Uploaded file deleted: {file_path}")
    except Exception as e:
        print(f"Error deleting uploaded files: {e}")

if __name__ == "__main__":
    app.run(debug=True, port=5001)
